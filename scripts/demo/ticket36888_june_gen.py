"""Демо-набор под сверку вкладки «Снятые баллы» с листом «июнь» (#36888).

Исходник — вложение «СВОД по чек листам с нарушениями по мес 2026.xlsx» задачи 36888
в учётной системе контура, лист «июнь». В репозитории этого файла нет и не будет: это
данные заказчика. Кто перегенерирует набор — берёт вложение из задачи, кладёт вне рабочей
копии и передаёт путь первым аргументом.

Выхлоп обезличен. Магазин, его адрес, город и формат заменяются синтетическими по порядку
первого появления в листе: один исходный город даёт один и тот же «Город NN», один формат —
один «Формат N». Регистр и лишние пробелы при этом схлопываются, поэтому «Санта Мини»
и «Санта мини» — один формат, а не два. Бланк проверки (разделы, пункты, их порядок и веса),
дата каждой проверки и снятые баллы по каждой ячейке переносятся дословно: по ним идёт
сверка, номер строки набора совпадает с номером строки листа. Таблица соответствия живёт
только в памяти процесса и никуда не пишется.

Раскладка листа не зашита: пункты — это колонки с весом в строке 6 и именем в строке 5,
раздел начинается там, где в строке 4 стоит непустой заголовок, данные идут до конца листа,
строки без номера ТО (протянутые формулы с #N/A) пропускаются.

Пишет два скрипта: заведение набора и парный к нему cleanup, снимающий набор со стенда.
Оба гоняются POST'ом тела на /eval/action работающего сервера и оба заканчиваются
EXPORT JSON с диагностикой: MESSAGE через эту ручку до вызывающего не доходит.

    python scripts/demo/ticket36888_june_gen.py <путь к xlsx> [--sheet июнь] [--limit 10]
"""
import argparse
import datetime
import os
import sys

import openpyxl
from openpyxl.utils import get_column_letter

TEMPLATE = "Проверка категории РЫБА — июнь 2026"
PREFIX = "DEMO36888"
COL_DATE, COL_SHOP, COL_CITY, COL_ADDRESS, COL_FORMAT = 1, 2, 3, 4, 5
ROW_SECTION, ROW_ITEM, ROW_WEIGHT = 4, 5, 6
DIMENSIONS = (("city", "Город"), ("format", "Формат"))


def fail(msg):
    sys.stderr.write(msg + "\n")
    sys.exit(2)


def lit(text):
    t = str(text).replace("\\", "\\\\")
    t = t.replace("'", "\\'")
    t = t.replace("{", "\\{").replace("}", "\\}")
    t = t.replace("$", "\\$")
    return "'" + t + "'"


def dec(x):
    return ("%.4f" % x).rstrip("0").rstrip(".") or "0"


def cell(row, col):
    return "%s%d" % (get_column_letter(col), row)


def num(value, addr, tolerated):
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip().replace(" ", "").replace(",", "."))
    except ValueError:
        if addr in tolerated:
            tolerated[addr] = value
            return None
        fail("ячейка %s: «%s» — не число. Пустая ячейка значит «баллы не сняты»; "
             "нечисловое значение того же значить не может. Excel такую ячейку в свою "
             "сумму тоже не берёт — если это разобрано и решено пропустить, скажите "
             "--tolerate %s." % (addr, value, addr))


class Ordinal:
    """Номер значения по порядку первого появления; регистр и пробелы не различаются."""

    def __init__(self):
        self.order = {}

    def of(self, value):
        key = " ".join(str(value).split()).casefold()
        if key not in self.order:
            self.order[key] = len(self.order) + 1
        return self.order[key]


def layout(ws):
    items = {}
    for c in range(1, ws.max_column + 1):
        weight = ws.cell(ROW_WEIGHT, c).value
        name = ws.cell(ROW_ITEM, c).value
        if not isinstance(weight, (int, float)):
            continue
        if name is None or not str(name).strip():
            fail("колонка %s: вес в строке %d есть, имени пункта в строке %d нет."
                 % (get_column_letter(c), ROW_WEIGHT, ROW_ITEM))
        items[c] = (str(name).strip(), float(weight), "r%02d" % c)
    if not items:
        fail("на листе не нашлось ни одного пункта: в строке %d нет весов." % ROW_WEIGHT)

    sections = []
    for c in sorted(items):
        title = ws.cell(ROW_SECTION, c).value
        title = str(title).strip() if title is not None and str(title).strip() else None
        if title:
            sections.append((title, []))
        elif not sections:
            fail("у первого пункта (колонка %s) нет раздела в строке %d."
                 % (get_column_letter(c), ROW_SECTION))
        sections[-1][1].append(c)
    return items, sections


def checks(ws, limit):
    rows = []
    for r in range(ROW_WEIGHT + 1, ws.max_row + 1):
        shop = ws.cell(r, COL_SHOP).value
        if shop is None or not str(shop).strip():
            continue
        date = ws.cell(r, COL_DATE).value
        if not isinstance(date, datetime.datetime):
            fail("строка %d: в %s ожидалась дата проверки, лежит «%s»."
                 % (r, cell(r, COL_DATE), date))
        rows.append((r, str(shop).strip(), date))
        if limit and len(rows) >= limit:
            break
    if not rows:
        fail("на листе нет ни одной строки-проверки: колонка %s пуста."
             % get_column_letter(COL_SHOP))
    return rows


def demo(ws, items, sections, rows, threshold, tolerated):
    assets, cities, formats = Ordinal(), Ordinal(), Ordinal()
    out = []
    w = out.append

    w("LOCAL tpl = StoreTask.Template();")
    w("LOCAL obj = StoreTask.CheckAsset();")
    w("")
    for did, dname in DIMENSIONS:
        w("IF NOT StoreTask.objectDimension(%s) THEN NEW d = StoreTask.ObjectDimension "
          "{ StoreTask.id(d) <- %s; StoreTask.name(d) <- %s; }"
          % (lit(did), lit(did), lit(dname)))
    w("")
    w("IF NOT StoreTask.templateByName(%s) THEN" % lit(TEMPLATE))
    w("    NEW t = StoreTask.Template {")
    w("        StoreTask.name(t) <- %s;" % lit(TEMPLATE))
    w("        StoreTask.passThreshold(t) <- %s;" % dec(threshold))
    for sname, cols in sections:
        w("        NEW s = StoreTask.Section { StoreTask.template(s) <- t; StoreTask.name(s) <- %s;"
          % lit(sname))
        for c in cols:
            name, weight, code = items[c]
            w("            NEW f = StoreTask.Field { StoreTask.section(f) <- s; StoreTask.code(f) <- %s; "
              "StoreTask.name(f) <- %s; StoreTask.type(f) <- StoreTask.FieldType.score; "
              "StoreTask.weight(f) <- %s; StoreTask.step(f) <- 0.5; StoreTask.scoreable(f) <- TRUE; }"
              % (lit(code), lit(name), dec(weight)))
        w("        }")
    w("    }")
    w("")
    w("tpl() <- StoreTask.templateByName(%s);" % lit(TEMPLATE))
    w("")

    for r, shop, date in rows:
        no = assets.of(shop)
        aid = "%s-%03d" % (PREFIX, no)
        city = ws.cell(r, COL_CITY).value
        fmt = ws.cell(r, COL_FORMAT).value
        city = "Город %02d" % cities.of(city) if city else None
        fmt = "Формат %d" % formats.of(fmt) if fmt else None
        address = ("%s, ул. Демонстрационная, %d" % (city, no) if city
                   else "ул. Демонстрационная, %d" % no)
        stamp = "%04d_%02d_%02d" % (date.year, date.month, date.day)

        w("    IF NOT StoreTask.checkAsset(%s) THEN" % lit(aid))
        w("        NEW ca = StoreTask.CheckAsset { StoreTask.assetId(ca) <- %s; StoreTask.assetName(ca) <- %s; }"
          % (lit(aid), lit("Объект %03d" % no)))
        w("    obj() <- StoreTask.checkAsset(%s);" % lit(aid))
        w("    IF NOT StoreTask.assetAddress(obj()) THEN StoreTask.assetAddress(obj()) <- %s;"
          % lit(address))
        for did, value in (("format", fmt), ("city", city)):
            if value:
                w("    IF NOT StoreTask.dimensionData(obj(), StoreTask.objectDimension(%s)) THEN "
                  "StoreTask.dimensionData(obj(), StoreTask.objectDimension(%s)) <- %s;"
                  % (lit(did), lit(did), lit(value)))
        w("    IF NOT (GROUP MAX StoreTask.Filling e2 IF StoreTask.template(e2) = tpl() "
          "AND StoreTask.checkObject(StoreTask.task(e2)) = obj() AND toDate(StoreTask.dateTime(e2)) = %s) THEN"
          % stamp)
        w("    NEW tk = StoreTask.Task {")
        w("        StoreTask.checkObject(tk) <- obj(); StoreTask.type(tk) <- StoreTask.checklistTaskType(); "
          "StoreTask.template(tk) <- tpl();")
        w("        NEW e = StoreTask.Filling {")
        w("            StoreTask.task(e) <- tk; StoreTask.template(e) <- tpl(); "
          "StoreTask.dateTime(e) <- toDateTime(%s);" % stamp)
        w("            StoreTask.numberValue(e, StoreTask.Field fl) <- StoreTask.weight(fl) "
          "WHERE StoreTask.template(fl) = tpl();")
        for c in sorted(items):
            v = num(ws.cell(r, c).value, cell(r, c), tolerated)
            if v is None or v <= 0:
                continue
            name, weight, code = items[c]
            w("            StoreTask.numberValue(e, StoreTask.field(tpl(), %s)) <- %s;"
              % (lit(code), dec(weight - min(v, weight))))
        w("            StoreTask.state(e) <- StoreTask.ExecutionState.finished;")
        w("        }")
        w("    }")

    like = lit(PREFIX + "-%")
    w("")
    w("APPLY;")
    w("")
    w("EXPORT JSON FROM")
    w("    failed = canceled(),")
    w("    msg = (OVERRIDE applyMessage(), ''),")
    w("    sections = (OVERRIDE (GROUP SUM 1 IF StoreTask.template(StoreTask.Section s2) = StoreTask.templateByName(%s)), 0),"
      % lit(TEMPLATE))
    w("    fields = (OVERRIDE (GROUP SUM 1 IF StoreTask.template(StoreTask.Field f2) = StoreTask.templateByName(%s)), 0),"
      % lit(TEMPLATE))
    w("    weightSum = (OVERRIDE (GROUP SUM StoreTask.weight(StoreTask.Field f3) IF StoreTask.template(f3) = StoreTask.templateByName(%s)), 0),"
      % lit(TEMPLATE))
    w("    assets = (OVERRIDE (GROUP SUM 1 IF StoreTask.assetId(StoreTask.CheckAsset a1) LIKE %s), 0)," % like)
    w("    withCity = (OVERRIDE (GROUP SUM 1 IF StoreTask.dimensionData(StoreTask.CheckAsset a2, StoreTask.objectDimension('city')) "
      "AND StoreTask.assetId(a2) LIKE %s), 0)," % like)
    w("    withFormat = (OVERRIDE (GROUP SUM 1 IF StoreTask.dimensionData(StoreTask.CheckAsset a3, StoreTask.objectDimension('format')) "
      "AND StoreTask.assetId(a3) LIKE %s), 0)," % like)
    w("    fillings = (OVERRIDE (GROUP SUM 1 IF StoreTask.template(StoreTask.Filling e3) = StoreTask.templateByName(%s)), 0),"
      % lit(TEMPLATE))
    w("    deducted = (OVERRIDE (GROUP SUM StoreTask.deductScore(StoreTask.Filling e4, StoreTask.Field f4) "
      "IF StoreTask.template(e4) = StoreTask.templateByName(%s) AND StoreTask.template(f4) = StoreTask.templateByName(%s)), 0);"
      % (lit(TEMPLATE), lit(TEMPLATE)))
    return out


def cleanup():
    like = lit(PREFIX + "-%")
    out = []
    w = out.append
    w("DELETE StoreTask.Task tk WHERE StoreTask.template(tk) = StoreTask.templateByName(%s);" % lit(TEMPLATE))
    w("DELETE StoreTask.Template t WHERE StoreTask.name(t) = %s;" % lit(TEMPLATE))
    w("DELETE StoreTask.CheckAsset ca WHERE StoreTask.assetId(ca) LIKE %s;" % like)
    w("")
    w("APPLY;")
    w("")
    w("EXPORT JSON FROM")
    w("    failed = canceled(),")
    w("    msg = (OVERRIDE applyMessage(), ''),")
    w("    templates = (OVERRIDE (GROUP SUM 1 IF StoreTask.name(StoreTask.Template t2) = %s), 0)," % lit(TEMPLATE))
    w("    assets = (OVERRIDE (GROUP SUM 1 IF StoreTask.assetId(StoreTask.CheckAsset a1) LIKE %s), 0)," % like)
    w("    tasks = (OVERRIDE (GROUP SUM 1 IF StoreTask.template(StoreTask.Task tk2) = StoreTask.templateByName(%s)), 0),"
      % lit(TEMPLATE))
    w("    fillings = (OVERRIDE (GROUP SUM 1 IF StoreTask.template(StoreTask.Filling e1) = StoreTask.templateByName(%s)), 0);"
      % lit(TEMPLATE))
    return out


def write(path, lines):
    with open(path, "w", encoding="utf-8", newline="\n") as f:
        f.write("\n".join(lines) + "\n")


def main():
    global TEMPLATE
    here = os.path.dirname(os.path.realpath(__file__))
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("xlsx", help="файл заказчика: вложение задачи 36888, вне рабочей копии")
    ap.add_argument("--sheet", default="июнь", help="лист свода (по умолчанию июнь)")
    ap.add_argument("--template", default=TEMPLATE, help="имя заводимого шаблона проверки")
    ap.add_argument("--threshold", type=float, default=80.0, help="проходной порог шаблона, %%")
    ap.add_argument("--out", default=os.path.join(here, "ticket36888_june_demo.lsf"))
    ap.add_argument("--cleanup-out", default=os.path.join(here, "ticket36888_june_cleanup.lsf"))
    ap.add_argument("--limit", type=int, default=0, help="взять только первые N строк-проверок")
    ap.add_argument("--tolerate", default="", help="ячейки с нечисловым значением, которые "
                    "разобраны и пропускаются: адреса через запятую, например AL78")
    args = ap.parse_args()

    TEMPLATE = args.template

    wb = openpyxl.load_workbook(args.xlsx, data_only=True)
    if args.sheet not in wb.sheetnames:
        fail("в книге нет листа «%s»; есть: %s." % (args.sheet, ", ".join(wb.sheetnames)))
    ws = wb[args.sheet]

    tolerated = dict.fromkeys(a.strip().upper() for a in args.tolerate.split(",") if a.strip())
    items, sections = layout(ws)
    rows = checks(ws, args.limit)
    write(args.out, demo(ws, items, sections, rows, args.threshold, tolerated))
    write(args.cleanup_out, cleanup())

    print("template:", TEMPLATE)
    print("sections:", len(sections), "items:", len(items),
          "weight sum:", dec(sum(i[1] for i in items.values())))
    for addr, value in sorted(tolerated.items()):
        print("пропущено по --tolerate:", addr, "=", "«%s»" % value if value is not None else "число, разбор не понадобился")
    print("rows:", len(rows), "->", args.out)
    print("cleanup ->", args.cleanup_out)


if __name__ == "__main__":
    main()
